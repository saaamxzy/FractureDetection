import os
import openpyxl
import pprint

data = {}

def record_data(wb):
  sheet = wb.get_sheet_by_name('Sheet1')
  for row in range(3, sheet.max_row+1):
    name    = "".join(sheet['A' + str(row)].value.split())
    sex     = "".join(sheet['B' + str(row)].value.split())
    age     = "".join(sheet['C' + str(row)].value.split())
    diag    = "".join(sheet['D' + str(row)].value.split())
    subj    = "".join(sheet['F' + str(row)].value.split())
    instrm  = "".join(sheet['R' + str(row)].value.split())
    conc    = "".join(sheet['Z' + str(row)].value.split())
    comm    = "".join(sheet['AA'+ str(row)].value.split())
    age_str = ""
    for d in age:
      if d.isdigit():
        age_str += d
      else:
        break
    age = int(age_str)
    data.setdefault(diag, {})
    data[diag]['name']  = name
    data[diag]['sex']   = sex
    data[diag]['age']   = age
    data[diag]['subj']  = subj
    data[diag]['instrm']= instrm
    data[diag]['conc']  = conc
    data[diag]['comm']  = comm

wb1 = openpyxl.load_workbook('1.xlsx')
record_data(wb1)
wb2 = openpyxl.load_workbook('2.xlsx')
record_data(wb2)
wb3 = openpyxl.load_workbook('3.xlsx')
record_data(wb3)
wb4 = openpyxl.load_workbook('4.xlsx')
record_data(wb4)

f = open('rs.txt', 'w')
f.write('allData :\n ' + pprint.pformat(data))
f.close()

filterData = {}

basedir = './'
curdir = basedir + 'data'
for fn in os.listdir(curdir):
  if not os.path.isdir(os.path.join(curdir, fn)):
    # not a directory
    continue
  fn_l = fn.split('_')
  key = fn_l[0]
  if not key.isdigit():
    continue
  value = data.get(key)
  if value != None:
    filterData[key] = value
  os.rename(os.path.join(curdir, fn),
            os.path.join('./final_data', key))

f = open('r.txt', 'w')
f.write('allData :\n ' + pprint.pformat(filterData))
f.close()

rswb = openpyxl.Workbook()
sht = rswb.get_sheet_by_name(rswb.get_sheet_names()[0])
sht['A1'] = 'diag_id'
sht['B1'] = 'name'
sht['C1'] = 'age'
sht['D1'] = 'sex'
sht['E1'] = '诊断项目'
sht['F1'] = '诊断仪器'
sht['G1'] = '诊断意见'
sht['H1'] = '诊断所见'

i = 2
for key, info in filterData.items():
  sht['A' + str(i)] = key
  sht['B' + str(i)] = info['name']
  sht['C' + str(i)] = info['age']
  sht['D' + str(i)] = info['sex']
  sht['E' + str(i)] = info['subj']
  sht['F' + str(i)] = info['instrm']
  sht['G' + str(i)] = info['conc']
  sht['H' + str(i)] = info['comm']

  i += 1

rswb.save('patients_info.xlsx')
